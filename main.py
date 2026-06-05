import os
import sys
import subprocess
import platform
import threading
import numpy as np
import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
from datetime import datetime
from openpyxl import Workbook, load_workbook

try:
    import cv2
    CV2_AVAILABLE = True
except Exception:
    CV2_AVAILABLE = False

try:
    import pyttsx3
    VOICE_AVAILABLE = True
except Exception:
    VOICE_AVAILABLE = False

# If OpenCV is missing, offer to install it automatically for the current Python.
if not CV2_AVAILABLE:
    try:
        tk.Tk().withdraw()
    except Exception:
        pass

    install = messagebox.askyesno(
        "Missing Dependency",
        "OpenCV (cv2) is not installed. Install opencv-contrib-python now?"
    )

    if install:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "opencv-contrib-python", "opencv-python"])
            # restart process to load newly installed packages
            os.execv(sys.executable, [sys.executable] + sys.argv)
        except Exception as e:
            messagebox.showerror("Install Failed", f"Automatic install failed: {e}\nPlease install manually: pip install opencv-contrib-python")
            sys.exit(1)
    else:
        messagebox.showerror("Missing Dependency", "OpenCV is required. Please install opencv-contrib-python and reopen the app.")
        sys.exit(1)

DATASET_DIR = "dataset"
TRAINER_DIR = "trainer"
TRAINER_FILE = "trainer/trainer.yml"
LABELS_FILE = "trainer/labels.txt"
STUDENTS_FILE = "students.xlsx"
ATTENDANCE_FILE = "attendance.xlsx"

ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "1234"

MAX_IMAGES = 120
CONFIDENCE_LIMIT = 35
REQUIRED_CONFIRMATIONS = 4

os.makedirs(DATASET_DIR, exist_ok=True)
os.makedirs(TRAINER_DIR, exist_ok=True)

status_label = None
login_status_label = None


def speak(text):
    if not VOICE_AVAILABLE:
        return

    def _tts(msg):
        try:
            engine = pyttsx3.init()
            engine.say(msg)
            engine.runAndWait()
        except Exception:
            pass

    try:
        t = threading.Thread(target=_tts, args=(text,), daemon=True)
        t.start()
    except Exception:
        pass


def create_excel_files():
    if not os.path.exists(STUDENTS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(["USN", "Name"])
        wb.save(STUDENTS_FILE)

    if not os.path.exists(ATTENDANCE_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["USN", "Name", "Date", "Time", "Status"])
        wb.save(ATTENDANCE_FILE)


def add_student(usn, name):
    create_excel_files()
    wb = load_workbook(STUDENTS_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == usn:
            return

    ws.append([usn, name])
    wb.save(STUDENTS_FILE)


def get_students():
    create_excel_files()
    wb = load_workbook(STUDENTS_FILE)
    ws = wb.active

    students = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            students[str(row[0])] = str(row[1])

    return students


def set_status(text):
    global status_label
    if status_label is not None:
        status_label.config(text=text)
        root.update_idletasks()


def open_camera_device():
    if platform.system() == "Windows":
        cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    else:
        cap = cv2.VideoCapture(0)

    if not cap.isOpened():
        return None

    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    cap.set(cv2.CAP_PROP_FPS, 30)
    if hasattr(cv2, 'CAP_PROP_BUFFERSIZE'):
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
    return cap


def register_face():
    set_status("Registering student face: enter details...")
    name = simpledialog.askstring("Register Student", "Enter student name:")
    if not name:
        set_status("Ready. Please choose an action.")
        return

    usn = simpledialog.askstring("Register Student", "Enter student USN:")
    if not usn:
        set_status("Ready. Please choose an action.")
        return

    name = name.strip()
    usn = usn.strip().upper().replace(" ", "_")

    add_student(usn, name)

    set_status(f"Saved {name}. Opening camera for registration...")

    detector = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

    cap = open_camera_device()
    if cap is None:
        messagebox.showerror("Camera Error", "Unable to open camera. Please check the device and try again.")
        set_status("Camera error. Ready.")
        return

    set_status("Capturing face images. Please look steadily at the camera.")

    count = 0

    # Use a scaled detection for speed; crop from original image for better quality
    detect_scale = 0.5

    while True:
        ret, frame = cap.read()
        if not ret:
            messagebox.showerror("Camera Error", "Camera not working")
            break

        frame = cv2.resize(frame, (640, 480))
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        small_gray = cv2.resize(gray, (0, 0), fx=detect_scale, fy=detect_scale)

        faces = detector.detectMultiScale(
            small_gray,
            scaleFactor=1.1,
            minNeighbors=5,
            minSize=(50, 50)
        )

        for (x, y, w, h) in faces:
            # map back to original coordinates
            x0 = int(x / detect_scale)
            y0 = int(y / detect_scale)
            w0 = int(w / detect_scale)
            h0 = int(h / detect_scale)

            count += 1
            face = gray[y0:y0+h0, x0:x0+w0]
            face = cv2.resize(face, (200, 200))

            cv2.imwrite(f"{DATASET_DIR}/{usn}_{count}.jpg", face)

            cv2.rectangle(frame, (x0, y0), (x0+w0, y0+h0), (0, 255, 0), 2)
            cv2.putText(frame, f"{count}/{MAX_IMAGES}", (20, 40),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

        cv2.imshow("Register Face - Press Q to Stop", frame)

        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

        if count >= MAX_IMAGES:
            break

    cap.release()
    cv2.destroyAllWindows()

    messagebox.showinfo("Success", f"Registration complete for {name} ({usn}).\nYou can now train the model.")
    set_status("Registration complete. Ready for next action.")


def train_model():
    set_status("Preparing training. Please wait...")
    try:
        recognizer = cv2.face.LBPHFaceRecognizer_create()
    except Exception:
        messagebox.showerror(
            "OpenCV Face Error",
            "Face recognizer is unavailable. Install opencv-contrib-python."
        )
        set_status("Train failed. OpenCV face module missing.")
        return

    faces = []
    labels = []
    label_map = {}
    current_id = 0

    for file_name in os.listdir(DATASET_DIR):
        if file_name.endswith(".jpg"):
            usn = file_name.split("_")[0]

            if usn not in label_map:
                label_map[usn] = current_id
                current_id += 1

            img_path = os.path.join(DATASET_DIR, file_name)
            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)

            if img is not None:
                img = cv2.resize(img, (200, 200))
                faces.append(img)
                labels.append(label_map[usn])

    if len(faces) == 0:
        messagebox.showwarning("No Data", "No face images found. Register student faces first.")
        set_status("No face data available. Register a student first.")
        return

    set_status("Training model... please wait.")

    recognizer.train(faces, np.array(labels))
    recognizer.save(TRAINER_FILE)

    with open(LABELS_FILE, "w") as f:
        for usn, label in label_map.items():
            f.write(f"{label},{usn}\n")

    messagebox.showinfo("Success", "Model trained successfully. You can now start attendance.")
    set_status("Model trained successfully. Ready for attendance.")


def load_labels():
    labels = {}

    if not os.path.exists(LABELS_FILE):
        return labels

    with open(LABELS_FILE, "r") as f:
        for line in f:
            label, usn = line.strip().split(",")
            labels[int(label)] = usn

    return labels


def mark_attendance(usn, name):
    create_excel_files()

    now = datetime.now()
    date = now.strftime("%d-%m-%Y")
    time = now.strftime("%I:%M:%S %p")

    wb = load_workbook(ATTENDANCE_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == usn and row[2] == date:
            return False
    ws.append([usn, name, date, time, "Present"])
    wb.save(ATTENDANCE_FILE)
    return True


def start_attendance():
    if not os.path.exists(TRAINER_FILE):
        messagebox.showwarning("Model Missing", "Train the model first.")
        set_status("Model missing. Train the model first.")
        return

    students = get_students()
    labels = load_labels()

    try:
        recognizer = cv2.face.LBPHFaceRecognizer_create()
    except Exception:
        messagebox.showerror(
            "OpenCV Face Error",
            "Face recognizer is unavailable. Install opencv-contrib-python."
        )
        set_status("Cannot start attendance. OpenCV face module missing.")
        return
    recognizer.read(TRAINER_FILE)

    detector = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
    cap = open_camera_device()
    if cap is None:
        messagebox.showerror("Camera Error", "Unable to open camera. Please check the device and try again.")
        set_status("Camera error. Ready.")
        return

    set_status("Starting attendance. Please look at the camera.")

    # initialize detection counters
    detection_count = {}

    # use scaled detection for speed
    detect_scale = 0.5

    while True:
        ret, frame = cap.read()
        if not ret:
            messagebox.showerror("Camera Error", "Camera not working")
            break

        frame = cv2.resize(frame, (640, 480))
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        small_gray = cv2.resize(gray, (0, 0), fx=detect_scale, fy=detect_scale)

        faces = detector.detectMultiScale(
            small_gray,
            scaleFactor=1.1,
            minNeighbors=5,
            minSize=(50, 50)
        )

        for (x, y, w, h) in faces:
            x0 = int(x / detect_scale)
            y0 = int(y / detect_scale)
            w0 = int(w / detect_scale)
            h0 = int(h / detect_scale)

            face = gray[y0:y0+h0, x0:x0+w0]
            face = cv2.resize(face, (200, 200))

            try:
                label, confidence = recognizer.predict(face)
            except Exception:
                continue

            if confidence < CONFIDENCE_LIMIT:
                usn = labels.get(label, "Unknown")
                if usn == "Unknown":
                    text = "Unknown"
                    color = (0, 0, 255)
                else:
                    name = students.get(usn, "Unknown")
                    detection_count[usn] = detection_count.get(usn, 0) + 1

                    if detection_count[usn] >= REQUIRED_CONFIRMATIONS:
                        saved = mark_attendance(usn, name)

                        if saved:
                            speak(f"Attendance marked for {name}")

                        text = f"{name} - Present"
                        color = (0, 255, 0)
                    else:
                        text = f"Verifying {name}"
                        color = (0, 255, 255)
            else:
                text = "Unknown"
                color = (0, 0, 255)

            cv2.rectangle(frame, (x0, y0), (x0+w0, y0+h0), color, 2)
            cv2.putText(frame, text, (x0, y0-10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)

        cv2.imshow("Face Attendance - Press Q to Stop", frame)

        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()
    set_status("Attendance ended. Ready for next action.")


def view_attendance():
    create_excel_files()

    selected_date = simpledialog.askstring(
        "Date Filter",
        "Enter date DD-MM-YYYY or leave empty for all:"
    )

    wb = load_workbook(ATTENDANCE_FILE)
    ws = wb.active

    view = tk.Toplevel(root)
    view.title("Attendance Records")
    view.geometry("850x450")
    view.configure(bg="#101820")

    tree = ttk.Treeview(
        view,
        columns=("USN", "Name", "Date", "Time", "Status"),
        show="headings"
    )

    for col in ("USN", "Name", "Date", "Time", "Status"):
        tree.heading(col, text=col)
        tree.column(col, width=150)

    tree.pack(fill="both", expand=True, padx=20, pady=20)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if selected_date and row[2] != selected_date:
            continue
        tree.insert("", tk.END, values=row)


def absent_report():
    create_excel_files()

    selected_date = simpledialog.askstring(
        "Absent Report",
        "Enter date DD-MM-YYYY:"
    )

    if not selected_date:
        return

    students = get_students()

    wb = load_workbook(ATTENDANCE_FILE)
    ws = wb.active

    present_usns = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == selected_date and row[4] == "Present":
            present_usns.add(row[0])

    view = tk.Toplevel(root)
    view.title("Absent Report")
    view.geometry("700x450")
    view.configure(bg="#101820")

    tree = ttk.Treeview(
        view,
        columns=("USN", "Name", "Date", "Status"),
        show="headings"
    )

    for col in ("USN", "Name", "Date", "Status"):
        tree.heading(col, text=col)
        tree.column(col, width=160)

    tree.pack(fill="both", expand=True, padx=20, pady=20)

    for usn, name in students.items():
        status = "Present" if usn in present_usns else "Absent"
        tree.insert("", tk.END, values=(usn, name, selected_date, status))


def open_excel():
    create_excel_files()

    file_path = os.path.abspath(ATTENDANCE_FILE)

    try:
        if platform.system() == "Windows":
            os.startfile(file_path)
        elif platform.system() == "Darwin":
            subprocess.call(["open", file_path])
        else:
            subprocess.call(["xdg-open", file_path])
    except:
        messagebox.showinfo("File Location", file_path)


def dashboard():
    for widget in root.winfo_children():
        widget.destroy()

    root.title("Face Attendance System")
    root.geometry("950x600")
    root.configure(bg="#101820")

    tk.Label(
        root,
        text="FACE ATTENDANCE SYSTEM",
        bg="#101820",
        fg="#00ffcc",
        font=("Arial", 28, "bold")
    ).pack(pady=25)

    tk.Label(
        root,
        text="Python + OpenCV + Excel + Admin Login",
        bg="#101820",
        fg="white",
        font=("Arial", 14)
    ).pack()

    frame = tk.Frame(root, bg="#101820")
    frame.pack(pady=35)

    style = {
        "width": 25,
        "height": 2,
        "font": ("Arial", 13, "bold"),
        "bg": "#00ffcc",
        "fg": "black"
    }

    tk.Button(frame, text="Register Student Face", command=register_face, **style).grid(row=0, column=0, padx=20, pady=12)
    tk.Button(frame, text="Train Model", command=train_model, **style).grid(row=0, column=1, padx=20, pady=12)
    tk.Button(frame, text="Start Attendance", command=start_attendance, **style).grid(row=1, column=0, padx=20, pady=12)
    tk.Button(frame, text="View Attendance", command=view_attendance, **style).grid(row=1, column=1, padx=20, pady=12)
    tk.Button(frame, text="Absent Report", command=absent_report, **style).grid(row=2, column=0, padx=20, pady=12)
    tk.Button(frame, text="Open Excel File", command=open_excel, **style).grid(row=2, column=1, padx=20, pady=12)

    global status_label
    status_label = tk.Label(
        root,
        text="Ready. Please choose an action.",
        bg="#101820",
        fg="#ffffff",
        font=("Arial", 14)
    )
    status_label.pack(pady=10)

    tk.Label(
        root,
        text="Strict Mode: attendance is marked only after strong repeated verification.",
        bg="#101820",
        fg="yellow",
        font=("Arial", 12)
    ).pack(pady=20)


def login():
    global login_status_label
    username = username_entry.get()
    password = password_entry.get()

    if login_status_label is not None:
        login_status_label.config(text="Processing... please wait")
        root.update_idletasks()

    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        if login_status_label is not None:
            login_status_label.config(text="Login successful. Opening dashboard...")
            root.update_idletasks()
        dashboard()
    else:
        if login_status_label is not None:
            login_status_label.config(text="Login failed. Check username and password.")
        messagebox.showerror("Login Failed", "Wrong username or password")


create_excel_files()

root = tk.Tk()
root.title("Admin Login")
root.geometry("600x400")
root.configure(bg="#101820")

tk.Label(
    root,
    text="ADMIN LOGIN",
    bg="#101820",
    fg="#00ffcc",
    font=("Arial", 28, "bold")
).pack(pady=40)

tk.Label(root, text="Enter admin username:", bg="#101820", fg="white", font=("Arial", 12)).pack()
username_entry = tk.Entry(root, font=("Arial", 16), width=25)
username_entry.pack(pady=6)

tk.Label(root, text="Enter admin password:", bg="#101820", fg="white", font=("Arial", 12)).pack()
password_entry = tk.Entry(root, font=("Arial", 16), width=25, show="*")
password_entry.pack(pady=6)

tk.Button(
    root,
    text="LOGIN",
    command=login,
    bg="#00ffcc",
    fg="black",
    font=("Arial", 15, "bold"),
    width=18
).pack(pady=25)

login_status_label = tk.Label(root, text="Enter username and password to login.", bg="#101820", fg="#ffffff", font=("Arial", 11))
login_status_label.pack()


root.mainloop()